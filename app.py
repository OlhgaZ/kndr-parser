"""
КНДР-парсер — Конкурентный анализ структуры страниц
Деплой: Streamlit Community Cloud
Загрузка страниц: ScrapingBee API (JS рендеринг, обход Cloudflare)
"""

import io
import re
import json
import string
import asyncio
import concurrent.futures
from urllib.parse import urlparse
from collections import defaultdict

import httpx
import streamlit as st
import pandas as pd
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
#  СЛОВАРЬ СИНОНИМОВ (русский + английский)
# ═══════════════════════════════════════════════════════════════════════════════
SYNONYM_GROUPS = {
    "🦸 Герой / Hero": [
        # RU
        "герой", "главный", "первый экран", "обложка", "баннер", "слайдер", "слайд",
        "добро пожаловать", "мы помогаем", "мы создаём", "мы делаем",
        # EN
        "hero", "welcome", "banner", "slider", "above the fold", "headline",
        "we help", "we create", "we build", "get started",
    ],
    "⭐ Преимущества": [
        "преимущество", "выгода", "польза", "плюс", "достоинство", "особенность",
        "почему мы", "почему нас", "зачем", "чем мы лучше", "наши преимущества",
        "why us", "why choose", "benefits", "advantages", "features", "what makes",
    ],
    "🛠️ Услуги / Продукты": [
        "услуга", "сервис", "решение", "продукт", "предложение", "направление",
        "что мы делаем", "что мы предлагаем", "наши услуги", "наши продукты",
        "services", "products", "solutions", "offerings", "what we do",
    ],
    "💰 Цены / Тарифы": [
        "цена", "стоимость", "тариф", "прайс", "расценка", "сколько стоит",
        "пакет", "тарифный план", "оплата", "стоить",
        "price", "pricing", "plans", "packages", "cost", "rates", "tariff",
    ],
    "💬 Отзывы": [
        "отзыв", "мнение", "оценка", "рецензия", "рекомендация", "что говорят",
        "наши клиенты", "нам доверяют",
        "testimonials", "reviews", "feedback", "what our clients say", "what people say",
    ],
    "❓ FAQ": [
        "faq", "вопрос", "ответ", "частый вопрос", "часто задают",
        "frequently asked", "questions", "q&a", "answers",
    ],
    "👥 Команда": [
        "команда", "сотрудник", "специалист", "эксперт", "мастер", "профессионал",
        "наша команда", "познакомьтесь",
        "team", "our team", "meet", "staff", "experts", "specialists",
    ],
    "🏢 О компании": [
        "компания", "организация", "студия", "агентство", "кто мы", "о нас",
        "история", "миссия", "ценности", "о компании",
        "about", "about us", "our story", "mission", "vision", "who we are",
    ],
    "📞 Контакты": [
        "контакт", "связь", "адрес", "телефон", "написать", "позвонить",
        "обратная связь", "свяжитесь", "форма связи",
        "contact", "contacts", "get in touch", "reach us", "write to us",
    ],
    "🗂️ Портфолио / Кейсы": [
        "портфолио", "кейс", "проект", "наши работы", "пример работ",
        "portfolio", "case studies", "our work", "projects", "examples",
    ],
    "🤝 Партнёры / Клиенты": [
        "партнёр", "партнер", "клиент", "логотип", "бренд", "нам доверяют",
        "с нами работают", "наши клиенты",
        "partners", "clients", "trusted by", "brands", "our clients",
    ],
    "🛡️ Гарантии": [
        "гарантия", "гарантировать", "обязательство", "уверенность", "надёжность",
        "guarantee", "guarantees", "warranty", "commitment", "reliability",
    ],
    "📋 Процесс / Этапы": [
        "процесс", "этап", "шаг", "как мы работаем", "порядок работы",
        "схема", "как это работает",
        "process", "how it works", "steps", "our process", "workflow",
    ],
    "📊 Статистика / Цифры": [
        "статистика", "цифра", "факт", "достижение", "результат", "показатель",
        "нас выбрали", "доверяют",
        "stats", "statistics", "numbers", "achievements", "results", "facts",
    ],
    "📝 Блог / Статьи": [
        "блог", "статья", "новость", "публикация", "пост", "материал",
        "blog", "articles", "news", "posts", "latest", "insights",
    ],
    "🎥 Видео": [
        "видео", "ролик", "презентация", "демо", "смотреть",
        "video", "watch", "demo", "presentation",
    ],
    "📩 CTA / Заявка": [
        "заявка", "записаться", "получить", "заказать", "попробовать",
        "начать", "купить", "оставить заявку", "бесплатно",
        "cta", "call to action", "sign up", "get started", "try free",
        "book", "order", "buy", "request",
    ],
}


def build_keyword_index() -> dict[str, str]:
    """Словарь keyword → group name"""
    idx = {}
    for group, keywords in SYNONYM_GROUPS.items():
        for kw in keywords:
            idx[kw.lower()] = group
    return idx


KEYWORD_INDEX = build_keyword_index()


def normalize(text: str) -> str:
    """Нормализация: lower + удаление пунктуации"""
    text = text.lower()
    text = re.sub(r"[^\w\s]", " ", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


def assign_group(heading: str) -> str:
    """Определяет семантическую группу заголовка по ключевым словам"""
    norm = normalize(heading)

    # Сначала ищем многословные фразы (длиннее — точнее)
    sorted_kws = sorted(KEYWORD_INDEX.keys(), key=len, reverse=True)
    for kw in sorted_kws:
        if kw in norm:
            return KEYWORD_INDEX[kw]

    # Fallback: первые значимые слова
    words = [w for w in norm.split() if len(w) > 3]
    return " ".join(words[:3]) if words else heading[:25]


# ═══════════════════════════════════════════════════════════════════════════════
#  ЗАГРУЗКА СТРАНИЦ (ScrapingBee)
# ═══════════════════════════════════════════════════════════════════════════════
def fetch_via_scrapingbee(url: str, api_key: str, timeout: int = 30) -> tuple[str | None, str | None]:
    """Загружает страницу через ScrapingBee API с JS-рендерингом"""
    endpoint = "https://app.scrapingbee.com/api/v1/"
    params = {
        "api_key": api_key,
        "url": url,
        "render_js": "true",
        "wait": "2000",           # ждём 2 сек после JS
        "wait_for": "body",
        "scroll_to_bottom": "true",
        "premium_proxy": "false",
        "block_ads": "true",
        "block_resources": "false",
        "return_page_source": "true",
    }
    try:
        with httpx.Client(timeout=timeout + 10) as client:
            r = client.get(endpoint, params=params)
        if r.status_code == 200:
            return r.text, None
        elif r.status_code == 401:
            return None, "❌ Неверный API-ключ ScrapingBee"
        elif r.status_code == 422:
            return None, f"❌ Сайт заблокировал парсинг (код 422)"
        elif r.status_code == 500:
            return None, "❌ ScrapingBee: внутренняя ошибка сервера"
        else:
            return None, f"❌ HTTP {r.status_code}: {r.text[:200]}"
    except httpx.TimeoutException:
        return None, f"❌ Таймаут ({timeout} сек)"
    except Exception as e:
        return None, f"❌ {str(e)[:150]}"


# ═══════════════════════════════════════════════════════════════════════════════
#  ПАРСИНГ HTML → БЛОКИ
# ═══════════════════════════════════════════════════════════════════════════════
def extract_blocks(html: str, mode: str) -> list[dict]:
    """
    Возвращает список блоков с метаданными.
    mode: 'main' | 'inner'
    """
    soup = BeautifulSoup(html, "lxml")

    # Убираем мусор
    for tag in soup(["script", "style", "noscript", "svg", "meta", "link"]):
        tag.decompose()

    body = soup.find("body") or soup

    if mode == "inner":
        for rem in body.find_all(["header", "footer", "nav"]):
            rem.decompose()

    search_root = body
    heading_tags = ["h1", "h2", "h3", "h4", "h5", "h6"]
    blocks = []
    seen_ids = set()

    for heading in search_root.find_all(heading_tags):
        heading_text = heading.get_text(strip=True)
        if not heading_text or len(heading_text) < 2:
            continue

        # Ближайший родительский section/div/article
        block_el = None
        current = heading.parent
        for _ in range(8):
            if current is None or current.name in ("body", "html", "[document]"):
                block_el = heading.parent
                break
            if current.name in ("section", "div", "article", "main", "aside"):
                block_el = current
                break
            current = current.parent
        if block_el is None:
            block_el = heading.parent

        el_id = id(block_el)
        if el_id in seen_ids:
            continue
        seen_ids.add(el_id)

        text = block_el.get_text(separator=" ", strip=True)

        # Кнопки: <button> + <a> с текстом
        buttons = block_el.find_all("button")
        links_cta = [
            a for a in block_el.find_all("a")
            if a.get_text(strip=True) and len(a.get_text(strip=True)) < 60
        ]
        forms = block_el.find_all("form")
        lists = block_el.find_all(["ul", "ol"])
        iframes = block_el.find_all("iframe")
        videos = block_el.find_all("video")
        tables = block_el.find_all("table")
        images = block_el.find_all("img")

        # FAQ Schema.org
        faq_schema = False
        for script in block_el.find_all("script", type="application/ld+json"):
            try:
                data = json.loads(script.string or "")
                items = [data] if isinstance(data, dict) else (data if isinstance(data, list) else [])
                for item in items:
                    if isinstance(item, dict) and item.get("@type") == "FAQPage":
                        faq_schema = True
            except Exception:
                pass

        blocks.append({
            "heading": heading_text,
            "level": int(heading.name[1]),
            "text_len": len(text),
            "buttons": len(buttons) + len(links_cta),
            "has_form": bool(forms),
            "has_list": bool(lists),
            "has_iframe": bool(iframes),
            "has_video": bool(videos),
            "has_table": bool(tables),
            "images": len(images),
            "has_faq_schema": faq_schema,
            "group": assign_group(heading_text),
        })

    return blocks


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
def make_excel(
    target_url: str,
    competitor_urls: list[str],
    all_results: dict[str, list[dict]],
) -> bytes:

    # Группируем блоки по group
    all_groups: dict[str, dict[str, list[dict]]] = {}
    for url, blocks in all_results.items():
        all_groups[url] = defaultdict(list)
        for b in blocks:
            all_groups[url][b["group"]].append(b)

    # Все уникальные группы, отсортированные
    unique_groups: list[str] = []
    seen = set()
    # Сначала в порядке из словаря (те что нашли)
    for g in SYNONYM_GROUPS.keys():
        for url in all_results:
            if g in all_groups.get(url, {}):
                if g not in seen:
                    unique_groups.append(g)
                    seen.add(g)
    # Затем неизвестные группы
    for url in all_results:
        for g in all_groups.get(url, {}):
            if g not in seen:
                unique_groups.append(g)
                seen.add(g)

    all_urls = competitor_urls + [target_url]
    short = {u: urlparse(u).netloc or u for u in all_urls}

    wb = openpyxl.Workbook()

    # ─── Стили ────────────────────────────────────────────────────────────────
    H_FILL   = PatternFill("solid", fgColor="1F3864")   # тёмно-синий заголовок
    TGT_FILL = PatternFill("solid", fgColor="D9EAD3")   # зелёный — анализируемый есть
    MISS_FILL= PatternFill("solid", fgColor="FCE4D6")   # красный — блок отсутствует
    OBL_FILL = PatternFill("solid", fgColor="C6EFCE")   # обязательно
    WISH_FILL= PatternFill("solid", fgColor="FFEB9C")   # желательно
    OPT_FILL = PatternFill("solid", fgColor="F2F2F2")   # по желанию
    ALT_FILL = PatternFill("solid", fgColor="EBF3FB")   # чередование строк (конкуренты)
    WHITE    = PatternFill("solid", fgColor="FFFFFF")

    H_FONT   = Font(bold=True, color="FFFFFF", size=10)
    BOLD     = Font(bold=True, size=10)
    NORM     = Font(size=10)
    CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    RIGHT    = Alignment(horizontal="right",  vertical="center")

    def border():
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)

    def set_header_row(ws, headers: list, row=1):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.fill = H_FILL; c.font = H_FONT
            c.alignment = CENTER; c.border = border()

    # ═══════════════════════════════════════════════════════════════════════════
    #  ЛИСТ 1 — Сравнение блоков
    # ═══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Сравнение блоков"

    comp_names = [short[u] for u in competitor_urls]
    cols = ["Блок"] + comp_names + [f"★ {short[target_url]}"] + ["Частота", "Рекомендация"]
    set_header_row(ws1, cols)

    for ri, group in enumerate(unique_groups, 2):
        freq = 0
        row_vals = [group]

        for cu in competitor_urls:
            gdata = all_groups.get(cu, {}).get(group)
            if gdata:
                freq += 1
                row_vals.append(f"✓  {gdata[0]['heading'][:45]}")
            else:
                row_vals.append("—")

        tgt_data = all_groups.get(target_url, {}).get(group)
        target_has = bool(tgt_data)
        if target_has:
            row_vals.append(f"✓  {tgt_data[0]['heading'][:45]}")
        else:
            row_vals.append("Отсутствует ✗")

        row_vals.append(freq)

        if freq >= 3:
            rec, rec_fill = "🔴 Обязательно", OBL_FILL
        elif freq == 2:
            rec, rec_fill = "🟡 Желательно", WISH_FILL
        else:
            rec, rec_fill = "⚪ По желанию", OPT_FILL
        row_vals.append(rec)

        row_bg = ALT_FILL if ri % 2 == 0 else WHITE

        for ci, val in enumerate(row_vals, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.border = border()
            c.alignment = LEFT if ci <= len(cols) - 2 else CENTER
            c.font = NORM

            if ci == 1:  # название группы
                c.font = BOLD
                c.fill = row_bg
            elif ci == len(cols) - 1:  # Частота
                c.alignment = CENTER
                c.fill = row_bg
            elif ci == len(cols):  # Рекомендация
                c.fill = rec_fill
                c.font = BOLD
                c.alignment = CENTER
            elif ci == len(cols) - 2:  # Колонка анализируемого
                c.fill = MISS_FILL if not target_has else TGT_FILL
            else:
                c.fill = row_bg

    # Ширина
    col_widths = [28] + [24] * len(competitor_urls) + [26, 10, 18]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[1].height = 32
    ws1.freeze_panes = "B2"
    ws1.sheet_view.showGridLines = True

    # ═══════════════════════════════════════════════════════════════════════════
    #  ЛИСТ 2 — Все заголовки
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Заголовки H1–H6")
    h2_cols = ["Сайт", "Уровень", "Заголовок", "Группа", "Длина текста",
               "CTA-кнопки", "Форма", "Список", "Изображения", "FAQ-схема"]
    set_header_row(ws2, h2_cols)

    ri2 = 2
    for url in all_urls:
        is_tgt = url == target_url
        for b in all_results.get(url, []):
            row = [
                short[url],
                f"H{b['level']}",
                b["heading"],
                b["group"],
                b["text_len"],
                b["buttons"],
                "Да" if b["has_form"] else "Нет",
                "Да" if b["has_list"] else "Нет",
                b["images"],
                "Да" if b["has_faq_schema"] else "Нет",
            ]
            row_bg = PatternFill("solid", fgColor="EBF3FB") if is_tgt else (ALT_FILL if ri2 % 2 == 0 else WHITE)
            for ci, val in enumerate(row, 1):
                c = ws2.cell(row=ri2, column=ci, value=val)
                c.border = border()
                c.font = BOLD if is_tgt else NORM
                c.fill = row_bg
                c.alignment = CENTER if ci in (2, 5, 6, 8, 9, 10) else LEFT
            ri2 += 1

    for i, w in enumerate([28, 9, 48, 26, 14, 12, 8, 8, 13, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[1].height = 32
    ws2.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════════════════════════════
    #  ЛИСТ 3 — Сводная статистика
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Сводная статистика")
    s_cols = ["Сайт", "Роль", "Блоков найдено", "CTA-кнопок", "Форм",
              "Списков", "Изображений", "Объём текста (симв.)", "FAQ-схем"]
    set_header_row(ws3, s_cols)

    for ri3, url in enumerate(all_urls, 2):
        blocks = all_results.get(url, [])
        is_tgt = url == target_url
        role = "★ Анализируемый" if is_tgt else "Конкурент"
        row = [
            short[url], role,
            len(blocks),
            sum(b["buttons"]   for b in blocks),
            sum(1 for b in blocks if b["has_form"]),
            sum(1 for b in blocks if b["has_list"]),
            sum(b["images"]    for b in blocks),
            sum(b["text_len"]  for b in blocks),
            sum(1 for b in blocks if b["has_faq_schema"]),
        ]
        fill = PatternFill("solid", fgColor="D9EAD3") if is_tgt else (ALT_FILL if ri3 % 2 == 0 else WHITE)
        for ci, val in enumerate(row, 1):
            c = ws3.cell(row=ri3, column=ci, value=val)
            c.border = border()
            c.font = BOLD if is_tgt else NORM
            c.fill = fill
            c.alignment = CENTER if ci > 2 else LEFT

    for i, w in enumerate([30, 18, 16, 14, 10, 10, 14, 22, 13], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.row_dimensions[1].height = 32
    ws3.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    st.set_page_config(
        page_title="КНДР-парсер",
        page_icon="🔍",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # ── CSS ──────────────────────────────────────────────────────────────────
    st.markdown("""
    <style>
    .main-title {
        font-size: 2.2rem; font-weight: 900;
        color: #1F3864; margin-bottom: 0; letter-spacing: -1px;
    }
    .subtitle {
        color: #555; margin-top: 2px; margin-bottom: 1.5rem; font-size: 1rem;
    }
    .metric-card {
        background: #f0f4f8; border-radius: 10px;
        padding: 12px 16px; text-align: center;
        border: 1px solid #d0dce8;
    }
    .status-ok  { color: #2e7d32; font-weight: 600; }
    .status-err { color: #c62828; font-weight: 600; }
    div[data-testid="stButton"] button[kind="primary"] {
        background: #1F3864; border: none;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown('<p class="main-title">🔍 КНДР-парсер</p>', unsafe_allow_html=True)
    st.markdown('<p class="subtitle">Конкурентный анализ структуры страниц · JS-рендеринг · Excel-отчёт</p>',
                unsafe_allow_html=True)

    # ── Боковая панель ────────────────────────────────────────────────────────
    with st.sidebar:
        st.header("⚙️ Настройки")

        api_key = st.text_input(
            "🔑 ScrapingBee API Key",
            type="password",
            help="Получите бесплатный ключ на scrapingbee.com (1000 запросов бесплатно)",
            placeholder="Вставьте API-ключ...",
        )
        if not api_key:
            st.info("👆 Введите API-ключ ScrapingBee для загрузки страниц.\n\n"
                    "[→ Получить бесплатный ключ](https://www.scrapingbee.com/)")

        st.divider()

        mode = st.radio(
            "Режим анализа",
            ["🏠 Главная страница", "📄 Внутренняя страница"],
            help="Главная: всё тело страницы.\nВнутренняя: без header и footer.",
        )
        mode_key = "main" if "Главная" in mode else "inner"

        timeout = st.slider("Таймаут на страницу (сек)", 15, 60, 30, step=5)

        st.divider()
        st.markdown("**Как работает:**")
        st.markdown("""
1. ScrapingBee загружает с JS-рендером
2. BeautifulSoup ищет блоки с H1–H6
3. Заголовки сопоставляются со словарём (~70 групп RU+EN)
4. Excel: сравнение · заголовки · статистика
""")
        with st.expander("📖 Группы блоков"):
            for group in SYNONYM_GROUPS:
                st.markdown(f"- {group}")

    # ── Ввод URL ──────────────────────────────────────────────────────────────
    col1, col2 = st.columns([1, 1], gap="large")

    with col1:
        st.subheader("🎯 Анализируемый сайт")
        target_url = st.text_input(
            "URL страницы",
            placeholder="https://mysite.ru",
            key="target_url",
        )

    with col2:
        st.subheader("🏆 Конкуренты")
        st.caption("От 4 до 10 URL, каждый с новой строки")
        competitors_raw = st.text_area(
            "URL конкурентов",
            placeholder=(
                "https://competitor1.ru\n"
                "https://competitor2.ru\n"
                "https://competitor3.ru\n"
                "https://competitor4.ru"
            ),
            height=170,
            key="competitors",
        )

    # Парсим конкурентов
    competitor_urls = [
        u.strip() for u in competitors_raw.strip().splitlines()
        if u.strip() and u.strip().startswith("http")
    ]

    # Валидация
    errors = []
    if target_url and not target_url.startswith("http"):
        errors.append("URL анализируемого сайта должен начинаться с http:// или https://")
    if competitor_urls and len(competitor_urls) < 4:
        errors.append(f"Нужно минимум 4 конкурента — сейчас введено: {len(competitor_urls)}")
    if competitor_urls and len(competitor_urls) > 10:
        errors.append(f"Максимум 10 конкурентов — сейчас введено: {len(competitor_urls)}")

    for e in errors:
        st.warning(f"⚠️ {e}")

    can_run = bool(api_key) and bool(target_url) and bool(competitor_urls) and not errors

    # ── Кнопки ───────────────────────────────────────────────────────────────
    st.divider()
    col_run, col_dl, col_info = st.columns([2, 2, 4])

    with col_run:
        run_btn = st.button(
            "🚀 Запустить анализ",
            disabled=not can_run,
            use_container_width=True,
            type="primary",
        )

    if not api_key:
        with col_info:
            st.info("🔑 Введите API-ключ ScrapingBee в боковой панели")
    elif not target_url or not competitor_urls:
        with col_info:
            st.info("⬆️ Заполните URL анализируемого сайта и конкурентов")

    # ── Анализ ────────────────────────────────────────────────────────────────
    if run_btn:
        all_urls = competitor_urls + [target_url]
        all_results: dict[str, list[dict]] = {}
        errors_log: dict[str, str] = {}

        progress_bar = st.progress(0)
        status_text  = st.empty()
        log_area     = st.container()

        total = len(all_urls)

        for idx, url in enumerate(all_urls):
            role = "анализируемый" if url == target_url else f"конкурент {idx + 1}/{len(competitor_urls)}"
            netloc = urlparse(url).netloc
            progress_bar.progress(idx / total, text=f"⏳ Загружаю {role}: {netloc}")
            status_text.info(f"Загружаю: **{url}**")

            html, err = fetch_via_scrapingbee(url, api_key, timeout)

            if err:
                errors_log[url] = err
                all_results[url] = []
                log_area.warning(f"⚠️ {netloc}: {err}")
            else:
                blocks = extract_blocks(html, mode_key)
                all_results[url] = blocks
                log_area.success(f"✅ {netloc} — найдено блоков: **{len(blocks)}**")

        progress_bar.progress(1.0, text="📊 Формирую Excel...")
        status_text.info("Создаю Excel-отчёт...")

        try:
            excel_bytes = make_excel(target_url, competitor_urls, all_results)
            st.session_state["excel_bytes"] = excel_bytes
            st.session_state["excel_ready"] = True
        except Exception as ex:
            st.error(f"Ошибка при создании Excel: {ex}")
            st.session_state["excel_ready"] = False

        progress_bar.progress(1.0, text="✅ Готово!")
        status_text.success("🎉 Анализ завершён!")

        # ── Превью ────────────────────────────────────────────────────────────
        st.divider()
        st.subheader("📋 Краткие результаты")

        cols = st.columns(min(len(all_urls), 6))
        for i, url in enumerate(all_urls):
            blocks = all_results.get(url, [])
            netloc = urlparse(url).netloc
            label  = f"★ {netloc}" if url == target_url else netloc
            with cols[i % len(cols)]:
                if url in errors_log:
                    st.metric(label, "Ошибка", delta="⚠️", delta_color="off")
                else:
                    groups_found = len(set(b["group"] for b in blocks))
                    st.metric(label, f"{len(blocks)} блоков", f"{groups_found} групп")

        # Список отсутствующих блоков
        target_blocks = all_results.get(target_url, [])
        target_groups = set(b["group"] for b in target_blocks)
        missing = []
        for cu in competitor_urls:
            for b in all_results.get(cu, []):
                if b["group"] not in target_groups:
                    missing.append(b["group"])

        from collections import Counter
        missing_freq = Counter(missing)
        if missing_freq:
            st.divider()
            st.subheader("🔴 Блоки, которых нет на вашем сайте")
            miss_df = pd.DataFrame(
                [(g, f, "🔴 Обязательно" if f >= 3 else "🟡 Желательно" if f == 2 else "⚪ По желанию")
                 for g, f in missing_freq.most_common()],
                columns=["Группа блока", "Частота у конкурентов", "Рекомендация"]
            )
            st.dataframe(miss_df, use_container_width=True, hide_index=True)

    # ── Кнопка скачивания ────────────────────────────────────────────────────
    if st.session_state.get("excel_ready"):
        with col_dl:
            st.download_button(
                label="📥 Скачать Excel",
                data=st.session_state["excel_bytes"],
                file_name="кндр_анализ.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )


if __name__ == "__main__":
    main()
